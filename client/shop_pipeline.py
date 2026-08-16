#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""店铺两步流水线：扫描写本地 md / 分析本地 md。

重新扫描 = module_scan_save_md（抓取结果 OCR 后写入 file/shops）
分析店铺 = module_analyze_shop_md（读 md → 词表/LLM → 由调用方上传 DB）
新店全自动 = 先扫描再分析。
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Callable

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

from shop_store import (  # noqa: E402
    find_local_shop_md,
    is_junk_product_title,
    overwrite_shop_md,
    parse_shop_md,
    save_shop_md_stats,
)


_NOISE_CTX_RE = re.compile(
    r"立即领取|最高立减|满减券|优惠券|补贴|币淘|领券|天猫币|淘金币|折上折"
)


def _llm_true_keywords(problem: str) -> set[str] | None:
    """有可解析 LLM JSON 时返回 violate=true 关键词；否则 None。"""
    jm = re.search(r"```json\s*(\{.*?\})\s*```", problem or "", re.S)
    if not jm:
        return None
    try:
        obj = json.loads(jm.group(1))
    except json.JSONDecodeError:
        return None
    if not isinstance(obj, dict) or "violations" not in obj:
        return None
    viols = obj.get("violations")
    if not isinstance(viols, list):
        return None
    true_kws: set[str] = set()
    for v in viols:
        if not isinstance(v, dict):
            continue
        val = v.get("violate")
        ok = val is True or val == 1
        if isinstance(val, str) and val.strip().lower() in ("true", "1", "yes"):
            ok = True
        if not ok:
            continue
        kw = str(v.get("keyword") or "").strip()
        if kw:
            true_kws.add(kw)
    return true_kws


def problem_to_row(item: dict) -> dict | None:
    """把扫描结果转成弹窗/Excel 一行；无问题返回 None。"""
    problem = str(item.get("problem") or "").strip()
    if not problem and not item.get("has_problem"):
        return None
    true_kws = _llm_true_keywords(problem)
    keywords: list[str] = []
    summaries: list[str] = []
    for m in re.finditer(r"命中「([^」]+)」[：:]?\s*(.*)$", problem, re.M):
        kw = m.group(1).strip()
        ctx = (m.group(2) or "").strip()
        if true_kws is not None and kw not in true_kws:
            continue
        if kw and kw not in keywords:
            keywords.append(kw)
        line_start = problem.rfind("\n", 0, m.start()) + 1
        line_end = problem.find("\n", m.end())
        if line_end < 0:
            line_end = len(problem)
        line = problem[line_start:line_end]
        src = ""
        sm = re.search(r"\[([^/\]]+)/([^\]]+)\]", line)
        if sm:
            src = sm.group(2).strip()
        if not ctx or _NOISE_CTX_RE.search(ctx):
            continue
        if src in ("主图文字", "标题"):
            label_src = "主图"
        elif src in ("详情文本", "详情图文字"):
            label_src = "详情页"
        else:
            label_src = src or ""
        label = (label_src + "：" if label_src else "") + ctx
        if len(label) > 80:
            label = label[:77] + "…"
        if label not in summaries:
            summaries.append(label)
    if true_kws is not None:
        for kw in sorted(true_kws):
            if kw not in keywords:
                keywords.append(kw)
        if not keywords and not summaries:
            return None
    if not keywords and not summaries and not problem:
        return None
    if keywords and not summaries:
        summaries = [f"命中「{kw}」" for kw in keywords]
    return {
        "goods_name": str(item.get("title") or item.get("id") or ""),
        "goods_link": str(item.get("url") or ""),
        "hit_keywords": "、".join(keywords) if keywords else "—",
        "hit_summary": "\n".join(summaries) if summaries else "—",
        "tb_item_id": str(item.get("id") or ""),
        "problem": problem,
    }


def module_scan_save_md(
    items: list[dict],
    shop: dict | None = None,
    *,
    overwrite: bool = False,
    do_ocr: bool = True,
    main_ocr_count: int = 2,
    detail_ocr_count: int = 6,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> dict:
    """扫描模块：对本机抓取结果做 OCR，写入 file/shops/<店名>.md。

    overwrite=True 表示「重新扫描」覆盖整店文件。
    不做词表/LLM，不上传数据库。
    """
    from local_scan import _ocr_item, probe_local_ocr

    shop = shop or {}
    shop_name = str(shop.get("shop_name") or "").strip() or "未命名店铺"
    shop_link = str(shop.get("shop_link") or shop.get("shop_url") or "").strip()
    if not shop_link and shop.get("shop_id"):
        shop_link = f"https://shop{shop['shop_id']}.taobao.com/"

    if do_ocr:
        ocr_st = probe_local_ocr()
        if ocr_st != "ok":
            raise RuntimeError(f"本机没有可用 OCR（不是滑块问题）: {ocr_st}")
        if progress_cb:
            progress_cb(0, max(len(items), 1), "本机 OCR 就绪")

    prepared: list[dict] = []
    total = len(items)
    for i, it in enumerate(items, 1):
        if not isinstance(it, dict):
            continue
        title = str(it.get("title") or "").strip()
        if title and is_junk_product_title(title):
            if progress_cb:
                progress_cb(i, total, f"跳过非商品「{title}」 {it.get('id') or ''}")
            continue
        if not it.get("ok", True) and not (it.get("title") or it.get("detail_texts")):
            continue
        work = dict(it)
        if do_ocr:
            if progress_cb:
                progress_cb(i, total, f"扫描 OCR {work.get('id') or i}")
            work = _ocr_item(work, main_ocr_count, detail_ocr_count, None)
        prepared.append(work)

    if not prepared:
        raise RuntimeError("没有可写入店铺文档的商品（抓取可能全部失败）")

    if overwrite:
        st = overwrite_shop_md(shop_name, prepared, shop_link=shop_link)
    else:
        st = save_shop_md_stats(shop_name, prepared, shop_link=shop_link)
    if progress_cb:
        progress_cb(
            total,
            total,
            f"本地店铺文档：新增 {st['added']}，跳过 {st['skipped']} → {st['path'].name}",
        )
    return {
        "ok": True,
        "shop_name": shop_name,
        "shop_link": shop_link,
        "path": st["path"],
        "added": st["added"],
        "skipped": st["skipped"],
        "items": prepared,
    }


def module_analyze_shop_md(
    shop_name: str,
    *,
    llm_conf: dict | None = None,
    do_llm: bool = True,
    max_items: int = 0,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> dict:
    """分析模块：读本地 shops md，对照词表/LLM，返回 scanned + problems（不上传）。

    max_items: >0 时只分析前 N 个商品；0=全部。
    """
    from local_scan import load_local_word_groups, purge_local_llm_secrets, scan_items_local

    path = find_local_shop_md(shop_name)
    parsed = parse_shop_md(path)
    items = list(parsed["items"] or [])
    if int(max_items or 0) > 0:
        items = items[: int(max_items)]

    purge_local_llm_secrets()
    groups = load_local_word_groups()
    if not any(w for _, w in groups):
        raise RuntimeError(
            "本地词表为空：请在界面保存极限词/错误描述后再扫"
        )
    use_llm = bool(do_llm and llm_conf)
    scanned = scan_items_local(
        items,
        groups,
        do_ocr=False,
        do_llm=use_llm,
        llm_conf=llm_conf if use_llm else None,
        progress_cb=progress_cb,
    )
    purge_local_llm_secrets()

    rows = []
    for it in scanned:
        row = problem_to_row(it)
        if row:
            rows.append(row)
    return {
        "ok": True,
        "shop_name": parsed.get("shop_name") or shop_name,
        "shop_link": parsed.get("shop_link") or "",
        "path": path,
        "items_from_md": items,
        "scanned": scanned,
        "problems": rows,
        "problem_count": len(rows),
        "analyze_limit": int(max_items or 0),
    }


def export_problems_xlsx(rows: list[dict], out_path: Path, shop_name: str = "") -> Path:
    """把问题列表写成 Excel（商品名 / 命中词 / 摘要 / 链接）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，请: python -m pip install openpyxl") from e

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "问题商品"
    headers = ["序号", "店铺", "商品名", "命中词", "摘要", "链接"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")
    for i, r in enumerate(rows, 1):
        ws.append(
            [
                i,
                shop_name,
                r.get("goods_name") or "",
                r.get("hit_keywords") or "",
                r.get("hit_summary") or "",
                r.get("goods_link") or "",
            ]
        )
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 45
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(out_path)
    return out_path


def export_pending_shops_xlsx(shops: list[dict], out_path: Path) -> Path:
    """待检查店铺表：店铺名 / 链接 / 淘宝店id。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，请: python -m pip install openpyxl") from e
    if not isinstance(shops, list):
        raise ValueError("shops 必须是列表")
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "待检查店铺"
    headers = ["序号", "店铺名", "链接", "淘宝店id"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")
    for i, s in enumerate(shops, 1):
        if not isinstance(s, dict):
            raise ValueError(f"第 {i} 条不是对象")
        ws.append(
            [
                i,
                s.get("shop_name") or "",
                s.get("shop_link") or s.get("url") or "",
                s.get("tb_shop_id") or s.get("shop_id") or "",
            ]
        )
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 16
    wb.save(out_path)
    return out_path
