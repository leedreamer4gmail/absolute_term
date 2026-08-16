#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""店铺商品极限词扫描器 v2 — 标题 + 详情文本 + 主图OCR + 详情图OCR 全覆盖。

用法:
    python3 scanner.py                    # 全量扫描, 输出 Excel
    python3 scanner.py --skip-ocr         # 跳过图片 OCR(只扫文本)
    python3 scanner.py --max-items 10     # 只扫前 N 个商品(测试)

配置(config.ini):
    [words]  limit_file / wrong_file = 词表文件(空格/换行分隔)
    [scan]   items_file / output_xlsx / ocr_url
    [llm]    复用根目录 llm_config.py, 可选 --llm 二次判定
"""
from __future__ import annotations

import argparse
import configparser
import csv
import json
import os
import re
import sys
import time
from pathlib import Path

import requests

HERE = Path(__file__).resolve().parent
CFG = HERE / "config.ini"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    Workbook = None

TIMEOUT = 30


def load_cfg() -> configparser.ConfigParser:
    cfg = configparser.ConfigParser()
    cfg.optionxform = str
    cfg.read(str(CFG), encoding="utf-8")
    return cfg


def _parse_words(text: str) -> list[str]:
    """词表按空白分隔（空格/换行/制表符），不再用 /。"""
    words: list[str] = []
    for piece in re.split(r"\s+", (text or "").strip()):
        piece = piece.strip()
        if piece and piece not in words:
            words.append(piece)
    return words


def _load_word_file(cfg: configparser.ConfigParser, option: str, fallback: str) -> list[str]:
    rel = cfg.get("words", option, fallback=fallback).strip()
    path = HERE / rel
    if not path.is_file():
        return []
    return _parse_words(path.read_text(encoding="utf-8"))


def load_word_groups(cfg: configparser.ConfigParser) -> list[tuple[str, list[str]]]:
    """加载极限词 + 错误描述两组词表。"""
    return [
        ("极限词", _load_word_file(cfg, "limit_file", "file/absolute_words.md")),
        ("错误描述", _load_word_file(cfg, "wrong_file", "file/wrong_word.md")),
    ]


def load_words(cfg: configparser.ConfigParser) -> list[str]:
    """合并两组词表（兼容旧调用）。"""
    words: list[str] = []
    for _label, group in load_word_groups(cfg):
        for w in group:
            if w not in words:
                words.append(w)
    if not words:
        sys.exit("词表为空：请检查 file/absolute_words.md 与 file/wrong_word.md")
    return words

def load_items(cfg: configparser.ConfigParser) -> list[dict]:
    items_file = cfg.get("scan", "items_file", fallback="data/items.csv")
    path = HERE / items_file
    if not path.is_file():
        sys.exit(f"商品数据文件不存在: {path}")
    items: list[dict] = []
    with path.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            items.append({
                "id": row.get("id", "").strip(),
                "title": row.get("title", "").strip(),
                "url": row.get("url", "").strip() or f"https://item.taobao.com/item.htm?id={row.get('id','').strip()}",
            })
    return items


def hit_context(text: str, keyword: str, radius: int = 10) -> str:
    idx = text.find(keyword)
    if idx < 0:
        return ""
    start = max(0, idx - radius)
    end = min(len(text), idx + len(keyword) + radius)
    prefix = "…" if start > 0 else ""
    suffix = "…" if end < len(text) else ""
    return f"{prefix}{text[start:end]}{suffix}"


def scan_texts(words: list[str], texts: dict[str, str], category: str = "") -> list[dict]:
    """对 {来源: 文本} 做词表扫描, 返回命中列表。"""
    hits: list[dict] = []
    for source, text in texts.items():
        if not text:
            continue
        for w in words:
            if w in text:
                hit = {
                    "source": source,
                    "keyword": w,
                    "context": hit_context(text, w),
                }
                if category:
                    hit["category"] = category
                hits.append(hit)
    return hits


def default_ocr_base(cfg: configparser.ConfigParser | None = None) -> str:
    if cfg is None:
        cfg = load_cfg()
    return cfg.get("ocr", "ocr_base", fallback="http://127.0.0.1:8799").rstrip("/")


def ocr_url(img_url: str, ocr_base: str = "") -> list[str]:
    """调服务器 OCR 服务识别图片文字。失败返回空。"""
    base = (ocr_base or default_ocr_base()).rstrip("/")
    try:
        r = requests.post(f"{base}/ocr_url", json={"url": img_url}, timeout=TIMEOUT + 20)
        r.raise_for_status()
        data = r.json()
        if data.get("error"):
            return []
        return data.get("texts", [])
    except Exception:  # noqa: BLE001
        return []


def local_images(item_id: str) -> list[Path]:
    """返回本地 images/item_{id}/ 下的图片(离线OCR用, 不依赖网络)。"""
    d = HERE / "images" / f"item_{item_id}"
    if not d.is_dir():
        return []
    return sorted(p for p in d.iterdir() if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp", ".gif"))


def ocr_local_file(img_path: Path, ocr_base: str = "") -> list[str]:
    """把本地图片 POST 到服务器 OCR(multipart)。失败返回空。"""
    base = (ocr_base or default_ocr_base()).rstrip("/")
    try:
        with img_path.open("rb") as f:
            r = requests.post(
                f"{base}/ocr",
                files={"file": (img_path.name, f, "image/jpeg")},
                timeout=TIMEOUT + 20,
            )
        r.raise_for_status()
        data = r.json()
        if data.get("error"):
            return []
        return data.get("texts", [])
    except Exception:  # noqa: BLE001
        return []


def _extract_json_obj(text: str) -> dict | None:
    """从模型输出中稳健提取 JSON 对象（避免贪婪正则把后续杂文吃进去）。"""
    raw = (text or "").strip()
    if not raw:
        return None
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.I)
    raw = re.sub(r"\s*```$", "", raw)
    try:
        obj = json.loads(raw)
        return obj if isinstance(obj, dict) else None
    except json.JSONDecodeError:
        pass
    decoder = json.JSONDecoder()
    for i, ch in enumerate(raw):
        if ch != "{":
            continue
        try:
            obj, _end = decoder.raw_decode(raw[i:])
            if isinstance(obj, dict):
                return obj
        except json.JSONDecodeError:
            continue
    return None


_DEFAULT_JUDGE_PROMPT = (
    "你是电商广告合规审核员。下面商品文本命中了极限词或错误描述候选, "
    "请判断哪些构成《广告法》绝对化用语违规或虚假宣传, 哪些是正常表述。\n"
    "判定原则：\n"
    "1. 只有宣称本品优于一切/无人能及/行业第一等「绝对化推销」才算违规，"
    "例如：最好、第一、天花板、最便宜、独一无二（吹效果）。\n"
    "2. 单字「最/第」等出现在客观规格、数量、顺序、时间里不算违规。"
    "例如：包装最小规格、最小起订量、最大承重、最新生产日期、"
    "最多可优惠40元、最后一件库存、最初配方、最高立减。\n"
    "3. 对每条命中都必须输出一条 violations，violate 填 true 或 false，并写简短 reason。\n\n"
    "商品标题: {title}\n链接: {url}\n\n{hits_block}\n{rules}"
    "只输出一个 JSON 对象，不要 markdown，不要解释。"
    '格式: {"violations":[{"keyword":"...","source":"...","violate":true,"reason":"..."}]}'
)


def _violate_is_true(val: object) -> bool:
    if val is True or val == 1:
        return True
    if isinstance(val, str) and val.strip().lower() in ("true", "1", "yes"):
        return True
    return False


def filter_hits_by_judge(hits: list[dict], judge: str) -> list[dict]:
    """LLM 已给出结构化判定时，只保留 violate=true 的词表命中；解析失败则原样返回。"""
    raw = (judge or "").strip()
    if not raw or not hits:
        return list(hits)
    obj = _extract_json_obj(raw)
    if not isinstance(obj, dict) or "violations" not in obj:
        return list(hits)
    viols = obj.get("violations")
    if not isinstance(viols, list):
        return list(hits)
    true_kws: set[str] = set()
    for v in viols:
        if not isinstance(v, dict):
            continue
        if not _violate_is_true(v.get("violate")):
            continue
        kw = str(v.get("keyword") or "").strip()
        if kw:
            true_kws.add(kw)
    return [h for h in hits if str(h.get("keyword") or "").strip() in true_kws]


def load_judge_prompt(config_path: str | Path | None = None, llm_conf: dict | None = None) -> str:
    """读取 LLM 判定提示词：优先 llm_conf['judge_prompt']，否则 config.ini [llm] judge_prompt。"""
    raw = ""
    if isinstance(llm_conf, dict):
        raw = str(llm_conf.get("judge_prompt") or "").strip()
    if not raw:
        import configparser

        cfg = configparser.ConfigParser()
        cfg.optionxform = str
        path = Path(config_path or CFG)
        if path.is_file():
            cfg.read(str(path), encoding="utf-8")
            local = path.with_name("config.ini.local")
            if local.is_file():
                cfg.read(str(local), encoding="utf-8")
            if cfg.has_option("llm", "judge_prompt"):
                raw = (cfg.get("llm", "judge_prompt") or "").strip()
    return raw or _DEFAULT_JUDGE_PROMPT


def apply_judge_prompt(template: str, *, title: str, url: str, hits_block: str, rules: str) -> str:
    """替换占位符；不用 str.format，避免模板里 JSON 花括号被误解析。"""
    out = template
    for key, val in (
        ("{title}", title),
        ("{url}", url),
        ("{hits_block}", hits_block),
        ("{rules}", rules),
    ):
        out = out.replace(key, val)
    return out


def llm_judge(
    item: dict,
    hits: list[dict],
    rules_text: str = "",
    config_path: str | Path | None = None,
    llm_conf: dict | None = None,
) -> str:
    """LLM 二次判定(可选): 汇总命中项判断是否真违规。返回结论文本。

    llm_conf: 内存中的 {api_url,model,api_key,judge_prompt?}（客户端从云端拉取，不落盘）。
    提示词来自 config.ini [llm] judge_prompt，勿硬编码业务文案。
    """
    from llm_config import call_chat

    hit_lines: list[str] = []
    for h in hits:
        cat = h.get("category") or "候选"
        hit_lines.append(
            f"- [{cat}/{h['source']}] 命中「{h['keyword']}」 上下文: {h['context']}"
        )
    hits_block = "\n".join(hit_lines)
    rules = f"《规则文档》\n{rules_text}\n\n" if rules_text else ""
    template = load_judge_prompt(config_path=config_path, llm_conf=llm_conf)
    prompt = apply_judge_prompt(
        template,
        title=str(item.get("title") or ""),
        url=str(item.get("url") or ""),
        hits_block=hits_block,
        rules=rules,
    )
    try:
        out = call_chat(
            [{"role": "user", "content": prompt}],
            temperature=0.2,
            timeout=90,
            config_path=None if llm_conf else (config_path or CFG),
            conf=llm_conf,
        )
        obj = _extract_json_obj(out)
        if obj is not None:
            return json.dumps(obj, ensure_ascii=False)
        return out[:500]
    except Exception as e:  # noqa: BLE001
        return f"LLM调用失败: {e}"


def build_problem_md(hits: list[dict], judge: str = "") -> str:
    """把命中结果写成 goods_tb.problem 用的 markdown。无问题返回空串。"""
    if not hits and not (judge or "").strip():
        return ""
    main_lines: list[str] = []
    detail_lines: list[str] = []
    other_lines: list[str] = []
    for h in hits:
        src = str(h.get("source") or "")
        cat = str(h.get("category") or "")
        kw = str(h.get("keyword") or "")
        ctx = str(h.get("context") or "")
        bullet = f"- [{cat}/{src}] 命中「{kw}」：{ctx}".rstrip("：")
        if src in ("主图文字", "标题"):
            main_lines.append(bullet)
        elif src in ("详情文本", "详情图文字"):
            detail_lines.append(bullet)
        else:
            other_lines.append(bullet)
    parts: list[str] = []
    if main_lines:
        parts.append("# 主图\n" + "\n".join(main_lines))
    if detail_lines:
        parts.append("# 详情\n" + "\n".join(detail_lines))
    if other_lines:
        parts.append("# 其它\n" + "\n".join(other_lines))
    j = (judge or "").strip()
    if j:
        parts.append("# LLM\n```json\n" + j + "\n```")
    return "\n\n".join(parts).strip()


_RAPID: object | None = None
_RAPID_ERR: str = ""


def _ocr_file_log(msg: str) -> None:
    path = (os.environ.get("ABSOLUTE_CLIENT_LOG") or "").strip()
    if not path:
        return
    try:
        with open(path, "a", encoding="utf-8") as f:
            f.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} {msg}\n")
    except OSError:
        pass


def _rapid() -> object:
    global _RAPID, _RAPID_ERR
    if _RAPID is None:
        try:
            from rapidocr_onnxruntime import RapidOCR

            _RAPID = RapidOCR()
            _RAPID_ERR = ""
        except Exception as e:  # noqa: BLE001
            _RAPID_ERR = str(e)
            _ocr_file_log(f"OCR_ENGINE_FAIL {e}")
            raise RuntimeError(f"本机 RapidOCR 不可用: {e}") from e
    return _RAPID


def probe_local_ocr() -> str:
    """探测本机 OCR。成功返回 ok，失败返回原因（不抛）。"""
    try:
        _rapid()
        return "ok"
    except Exception as e:  # noqa: BLE001
        return str(e)


def ocr_image_bytes(data: bytes) -> list[str]:
    """本机 RapidOCR 识别图片字节，返回文本行。"""
    if not data:
        return []
    import io
    import tempfile

    from PIL import Image

    try:
        im = Image.open(io.BytesIO(data))
        im.load()
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"图片无法解码: {e}") from e
    if im.mode not in ("RGB", "L"):
        im = im.convert("RGB")
    elif im.mode == "L":
        im = im.convert("RGB")
    engine = _rapid()
    path = ""
    try:
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            im.save(tmp, format="JPEG", quality=95)
            path = tmp.name
        result, _ = engine(path)
    finally:
        if path:
            try:
                Path(path).unlink(missing_ok=True)
            except OSError:
                pass
    if not result:
        return []
    return [str(line[1]).strip() for line in result if line and len(line) > 1 and str(line[1]).strip()]


def ocr_image_url_local(img_url: str, timeout: int = 45) -> list[str]:
    """下载图片后本机 OCR。带淘宝 Cookie/Referer；失败抛错，不假装没字。"""
    url = (img_url or "").strip()
    if not url:
        return []
    if url.startswith("//"):
        url = "https:" + url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        ),
        "Referer": "https://item.taobao.com/",
        "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
    }
    try:
        from fetch_item import load_cookie

        ck = load_cookie()
        if ck:
            headers["Cookie"] = ck
    except Exception:  # noqa: BLE001
        pass
    try:
        r = requests.get(url, timeout=timeout, headers=headers)
        r.raise_for_status()
        ctype = (r.headers.get("Content-Type") or "").lower()
        if "text/html" in ctype or r.content[:15].lstrip().startswith(b"<!DOCTYPE"):
            raise RuntimeError(f"下载到的不是图片（Content-Type={ctype or '空'}）")
        return ocr_image_bytes(r.content)
    except Exception as e:  # noqa: BLE001
        _ocr_file_log(f"OCR_FAIL url={url[:160]} err={e}")
        raise RuntimeError(f"OCR 下载/识别失败: {e}") from e


def write_xlsx(results: list[dict], out_path: Path):
    if Workbook is None:  # pragma: no cover
        raise RuntimeError("缺少 openpyxl, 请: .venv-abs/bin/pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "违规清单"
    headers = ["序号", "店铺/商品ID", "标题", "链接", "违规来源", "命中词", "上下文", "详情页OCR文本", "判定"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="FFC7CE")
    for i, r in enumerate(results, 1):
        row = [
            i,
            r["id"],
            r["title"],
            r["url"],
            "、".join({h["source"] for h in r["hits"]}),
            "、".join({h["keyword"] for h in r["hits"]}),
            " || ".join({h["context"] for h in r["hits"]}),
            r["ocr_summary"],
            r.get("judge", ""),
        ]
        ws.append(row)
        if r["hits"]:
            for c in ws[ws.max_row]:
                c.fill = red
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 60
    ws.column_dimensions["H"].width = 80
    ws.column_dimensions["I"].width = 50
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(out_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="店铺商品极限词扫描 v2")
    parser.add_argument("--skip-ocr", action="store_true", help="跳过图片 OCR")
    parser.add_argument("--max-items", type=int, default=0, help="只扫前 N 个商品")
    parser.add_argument("--llm", action="store_true", help="LLM 二次判定")
    args = parser.parse_args()

    cfg = load_cfg()
    word_groups = load_word_groups(cfg)
    words = load_words(cfg)
    items = load_items(cfg)
    if args.max_items:
        items = items[: args.max_items]
    ocr_base = cfg.get("ocr", "ocr_base", fallback=default_ocr_base(cfg))

    for label, group in word_groups:
        print(f"{label}({len(group)}): {group}")
    print(f"商品数: {len(items)}  OCR: {'跳过' if args.skip_ocr else ocr_base}")

    results: list[dict] = []
    total_hits = 0
    for n, it in enumerate(items, 1):
        print(f"[{n}/{len(items)}] {it['id']} {it['title'][:30]}...")
        texts: dict[str, str] = {"标题": it["title"]}
        ocr_texts: list[str] = []

        # 抓详情(标题/详情文本/图URL) — 无 cookie 时跳过详情, 只用标题
        from fetch_item import fetch_item

        detail = fetch_item(it["id"])
        if detail.get("title"):
            texts["标题"] = detail["title"]
        if detail.get("detail_texts"):
            texts["详情文本"] = "\n".join(detail["detail_texts"])
        if detail.get("error"):
            print(f"  ⚠ {detail['error']}")

        # 图片 OCR: 优先本地已下载图, 否则主图/详情图URL
        if not args.skip_ocr:
            local = local_images(it["id"])
            if local:
                for p in local:
                    t = ocr_local_file(p, ocr_base)
                    if t:
                        ocr_texts.extend(t)
                        print(f"  OCR[本地] {p.name[:40]}: {len(t)} 行")
            else:
                all_imgs = detail.get("main_image_urls", [])[:4] + detail.get("detail_image_urls", [])[:12]
                for u in all_imgs:
                    t = ocr_url(u, ocr_base)
                    if t:
                        ocr_texts.extend(t)
                        print(f"  OCR {u.rsplit('/', 1)[-1][:40]}: {len(t)} 行")
            if ocr_texts:
                texts["图片文字"] = "\n".join(ocr_texts)

        hits: list[dict] = []
        for category, group in word_groups:
            if group:
                hits.extend(scan_texts(group, texts, category=category))
        # 去重
        seen = set()
        uniq_hits = []
        for h in hits:
            k = (h.get("category", ""), h["source"], h["keyword"])
            if k not in seen:
                seen.add(k)
                uniq_hits.append(h)
        if uniq_hits:
            total_hits += len(uniq_hits)
        judge = ""
        if args.llm and uniq_hits:
            rules = ""
            rules_file = HERE / "rules.md"
            if rules_file.is_file():
                rules = rules_file.read_text(encoding="utf-8")
            judge = llm_judge(it, uniq_hits, rules)
            print(f"  LLM: {judge[:150]}")

        results.append({
            "id": it["id"],
            "title": texts["标题"],
            "url": it["url"],
            "hits": uniq_hits,
            "ocr_summary": "\n".join(ocr_texts)[:500],
            "judge": judge,
        })

    out_path = HERE / cfg.get("scan", "output_xlsx", fallback="output/违规清单.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_xlsx(results, out_path)

    n_violated = sum(1 for r in results if r["hits"])
    print(f"\n完成: {len(results)} 个商品, {n_violated} 个有命中, 共 {total_hits} 处")
    print(f"Excel: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
